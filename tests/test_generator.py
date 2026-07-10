import json
import math
import subprocess
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "make_evidence_charting_workbook.py"


def test_self_test_workbook_contract(tmp_path):
    output = tmp_path / "self-test.xlsx"
    result = subprocess.run([sys.executable, str(SCRIPT), "--self-test", "--output", str(output)], capture_output=True, text=True, check=True)
    summary = json.loads(result.stdout)
    assert summary["blank_cells"] == 1
    assert summary["zero_cells"] == 0
    assert summary["embedded_images"] == 10
    with zipfile.ZipFile(output) as package:
        assert not any(name.startswith("xl/tables/") for name in package.namelist())
    workbook = openpyxl.load_workbook(output, data_only=True)
    assert workbook.sheetnames == ["00_阅读说明", "01_总览", "02_所有柱状图", "03_差距对比图", "04_按领域", "05_颜色标准", "06_缺失值审计", "07_差距摘要", "08_原始数据", "09_来源"]
    latency = next(row for row in workbook["07_差距摘要"].iter_rows(min_row=2, values_only=True) if row[1] == "Response time")
    assert math.isclose(latency[10], 13.265306, rel_tol=1e-5)
    workbook.close()


def test_validate_only_rejects_string_boolean(tmp_path):
    spec = {"entities": [{"id": "a"}, {"id": "b"}], "metrics": [{"name": "m", "higher_is_better": "false", "scores": {"a": 1, "b": 2}, "sources": ["user-provided data"]}]}
    path = tmp_path / "bad.json"
    path.write_text(json.dumps(spec), encoding="utf-8")
    result = subprocess.run([sys.executable, str(SCRIPT), "--validate-only", "--input", str(path)], capture_output=True, text=True)
    assert result.returncode != 0
    assert "must be true or false" in result.stderr


def test_single_score_is_not_counted_as_a_win():
    import importlib.util

    module_spec = importlib.util.spec_from_file_location("generator", SCRIPT)
    module = importlib.util.module_from_spec(module_spec)
    module_spec.loader.exec_module(module)
    spec = module.normalize_spec({"entities": [{"id": "a"}, {"id": "b"}], "metrics": [{"name": "coverage", "scores": {"a": 1}, "sources": ["user-provided data"]}]})
    stats = module.compute_model_stats(spec)
    assert stats["win_counts"] == {}
    assert stats["summary_rows"][0]["winner"] == ""


def test_long_text_is_wrapped_sized_and_preserved(tmp_path):
    long_note = "这是一段用于验证中文自动换行和行高估算的证据说明。" * 45
    spec = {
        "entities": [{"id": "a", "label": "模型甲"}, {"id": "b", "label": "Model B"}],
        "metrics": [{"name": "Long note", "scores": {"a": 1, "b": 2}, "sources": ["user-provided data"], "evidence_notes": long_note}],
    }
    input_path = tmp_path / "long.json"
    output_path = tmp_path / "long.xlsx"
    input_path.write_text(json.dumps(spec, ensure_ascii=False), encoding="utf-8")
    subprocess.run([sys.executable, str(SCRIPT), "--input", str(input_path), "--output", str(output_path)], check=True)
    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook["09_来源"]
    note_cell = sheet.cell(2, 10)
    assert note_cell.alignment.wrap_text is True
    assert note_cell.comment.text == long_note
    assert "see comment" in note_cell.value
    assert 24 < sheet.row_dimensions[2].height <= 180
    assert sheet.column_dimensions["J"].width <= 48
    workbook.close()
