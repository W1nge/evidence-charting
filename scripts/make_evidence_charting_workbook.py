#!/usr/bin/env python
"""Generate a source-backed bar-chart research workbook from a JSON spec."""

from __future__ import annotations

import argparse
import json
import math
import shutil
import sys
import tempfile
import textwrap
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any


DEFAULT_PALETTE = [
    "#0B5CAD",
    "#159A8C",
    "#34A853",
    "#6C3BAA",
    "#8E8E93",
    "#C2410C",
    "#F97316",
    "#B45309",
    "#92400E",
    "#4285F4",
    "#0F9D58",
    "#7C3AED",
    "#DB2777",
    "#0891B2",
]


def require_dependencies() -> None:
    missing = []
    for module in ["openpyxl", "matplotlib", "numpy"]:
        try:
            __import__(module)
        except Exception:
            missing.append(module)
    if missing:
        joined = ", ".join(missing)
        raise SystemExit(f"Missing required Python packages: {joined}")


def setup_matplotlib():
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib import font_manager

    for fp in [
        r"C:\Windows\Fonts\msyh.ttc",
        r"C:\Windows\Fonts\simhei.ttf",
        r"C:\Windows\Fonts\simsun.ttc",
    ]:
        path = Path(fp)
        if path.exists():
            font_manager.fontManager.addfont(str(path))
            prop = font_manager.FontProperties(fname=str(path))
            plt.rcParams["font.family"] = prop.get_name()
            break
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["figure.dpi"] = 150
    return plt


def normalize_color(value: str | None, index: int) -> str:
    if not value:
        return DEFAULT_PALETTE[index % len(DEFAULT_PALETTE)]
    value = str(value).strip()
    if not value.startswith("#"):
        value = "#" + value
    if len(value) != 7:
        return DEFAULT_PALETTE[index % len(DEFAULT_PALETTE)]
    return value.upper()


def sheet_color(hex_color: str) -> str:
    return hex_color.replace("#", "").upper()


def as_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        return float(str(value).replace(",", ""))
    except Exception:
        return None


def load_spec(path: Path | None, self_test: bool) -> dict[str, Any]:
    if self_test:
        return {
            "title": "Self-test benchmark research",
            "generated_at": datetime.now().strftime("%Y-%m-%d"),
            "source_note": "Synthetic self-test data.",
            "entities": [
                {"id": "alpha", "label": "Alpha 1", "group": "Target", "color": "#0B5CAD"},
                {"id": "beta", "label": "Beta 1", "group": "Target", "color": "#159A8C"},
                {"id": "base", "label": "Baseline", "group": "Baseline", "color": "#8E8E93"},
                {"id": "rival", "label": "Rival", "group": "External", "color": "#C2410C"},
            ],
            "metrics": [
                {
                    "category": "Coding",
                    "name": "Task A",
                    "unit": "%",
                    "higher_is_better": True,
                    "scores": {"alpha": 74.2, "beta": 71.5, "base": 68.0, "rival": 72.0},
                    "sources": ["self-test"],
                    "evidence_notes": "Synthetic primary-like test row.",
                },
                {
                    "category": "Coding",
                    "name": "Task B",
                    "unit": "%",
                    "higher_is_better": True,
                    "scores": {"alpha": 81.0, "beta": 79.2, "base": 75.5},
                    "sources": ["self-test"],
                    "evidence_notes": "Synthetic row with one missing rival score.",
                },
                {
                    "category": "Latency",
                    "name": "Response time",
                    "unit": "ms",
                    "higher_is_better": False,
                    "scores": {"alpha": 920, "beta": 850, "base": 980, "rival": 890},
                    "sources": ["self-test"],
                    "evidence_notes": "Synthetic lower-is-better metric.",
                },
            ],
            "comparison": {
                "primary_group": "Target",
                "baseline_entity": "base",
                "external_groups": ["External"],
            },
        }
    if not path:
        raise SystemExit("Provide --input data.json or --self-test.")
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def normalize_spec(spec: dict[str, Any]) -> dict[str, Any]:
    entities = spec.get("entities") or []
    metrics = spec.get("metrics") or []
    if len(entities) < 2:
        raise SystemExit("Spec must include at least two entities.")
    if not metrics:
        raise SystemExit("Spec must include at least one metric.")

    seen_ids = set()
    normalized_entities = []
    for idx, entity in enumerate(entities):
        eid = str(entity.get("id") or "").strip()
        label = str(entity.get("label") or eid).strip()
        if not eid:
            raise SystemExit(f"Entity at index {idx} is missing id.")
        if eid in seen_ids:
            raise SystemExit(f"Duplicate entity id: {eid}")
        seen_ids.add(eid)
        normalized_entities.append(
            {
                "id": eid,
                "label": label,
                "group": str(entity.get("group") or "Entities").strip(),
                "color": normalize_color(entity.get("color"), idx),
            }
        )

    normalized_metrics = []
    for idx, metric in enumerate(metrics):
        name = str(metric.get("name") or f"Metric {idx + 1}").strip()
        scores = metric.get("scores") or {}
        normalized_scores = {e["id"]: as_number(scores.get(e["id"])) for e in normalized_entities}
        if not any(v is not None for v in normalized_scores.values()):
            raise SystemExit(f"Metric has no numeric scores: {name}")
        sources = metric.get("sources") or metric.get("source") or []
        if isinstance(sources, str):
            sources = [sources]
        evidence_notes = str(metric.get("evidence_notes") or metric.get("evidence_note") or "").strip()
        normalized_metrics.append(
            {
                "category": str(metric.get("category") or "General").strip(),
                "name": name,
                "unit": str(metric.get("unit") or "Score").strip(),
                "higher_is_better": bool(metric.get("higher_is_better", True)),
                "scores": normalized_scores,
                "sources": [str(s) for s in sources if s],
                "evidence_notes": evidence_notes,
            }
        )

    spec["entities"] = normalized_entities
    spec["metrics"] = normalized_metrics
    spec["title"] = str(spec.get("title") or "Bar chart research report")
    spec["generated_at"] = str(spec.get("generated_at") or datetime.now().strftime("%Y-%m-%d"))
    spec["source_note"] = str(spec.get("source_note") or "")
    spec["comparison"] = spec.get("comparison") or {}
    return spec


def best_entity(candidates: list[tuple[str, float]], higher_is_better: bool) -> tuple[str, float] | None:
    if not candidates:
        return None
    return max(candidates, key=lambda x: x[1]) if higher_is_better else min(candidates, key=lambda x: x[1])


def compute_model_stats(spec: dict[str, Any]) -> dict[str, Any]:
    entities = spec["entities"]
    entity_ids = [e["id"] for e in entities]
    id_to_entity = {e["id"]: e for e in entities}
    metrics = spec["metrics"]
    comparison = spec.get("comparison") or {}
    primary_group = comparison.get("primary_group")
    baseline_entity = comparison.get("baseline_entity")
    external_groups = set(comparison.get("external_groups") or [])

    coverage = {eid: 0 for eid in entity_ids}
    zero_count = 0
    numeric_count = 0
    win_counts = defaultdict(int)
    best_primary_counts = defaultdict(int)
    best_external_counts = defaultdict(int)
    summary_rows = []

    for metric in metrics:
        higher = metric["higher_is_better"]
        values = [(eid, metric["scores"].get(eid)) for eid in entity_ids]
        numeric_values = [(eid, value) for eid, value in values if isinstance(value, (int, float))]
        for eid, value in numeric_values:
            coverage[eid] += 1
            numeric_count += 1
            if value == 0:
                zero_count += 1

        winning = best_entity(numeric_values, higher)
        winners = []
        if winning:
            best_value = winning[1]
            winners = [eid for eid, value in numeric_values if value == best_value]
            for eid in winners:
                win_counts[eid] += 1

        primary_candidates = []
        if primary_group:
            primary_candidates = [
                (eid, value)
                for eid, value in numeric_values
                if id_to_entity[eid]["group"] == primary_group
            ]
        best_primary = best_entity(primary_candidates, higher) if primary_candidates else None
        if best_primary:
            best_primary_counts[best_primary[0]] += 1

        baseline_value = metric["scores"].get(baseline_entity) if baseline_entity else None

        external_candidates = []
        if external_groups:
            external_candidates = [
                (eid, value)
                for eid, value in numeric_values
                if id_to_entity[eid]["group"] in external_groups
            ]
        elif primary_group:
            external_candidates = [
                (eid, value)
                for eid, value in numeric_values
                if id_to_entity[eid]["group"] != primary_group and eid != baseline_entity
            ]
        best_external = best_entity(external_candidates, higher) if external_candidates else None
        if best_external:
            best_external_counts[best_external[0]] += 1

        delta_baseline = None
        delta_baseline_pct = None
        if best_primary and isinstance(baseline_value, (int, float)):
            delta_baseline = best_primary[1] - baseline_value
            if not higher:
                delta_baseline = -delta_baseline
            if baseline_value != 0:
                raw_delta = best_primary[1] - baseline_value
                delta_baseline_pct = raw_delta / baseline_value * 100

        delta_external = None
        delta_external_pct = None
        if best_primary and best_external:
            delta_external = best_primary[1] - best_external[1]
            if not higher:
                delta_external = -delta_external
            if best_external[1] != 0:
                raw_delta = best_primary[1] - best_external[1]
                delta_external_pct = raw_delta / best_external[1] * 100

        summary_rows.append(
            {
                "category": metric["category"],
                "metric": metric["name"],
                "unit": metric["unit"],
                "higher_is_better": higher,
                "winner": ", ".join(id_to_entity[eid]["label"] for eid in winners),
                "best_primary_id": best_primary[0] if best_primary else None,
                "best_primary_score": best_primary[1] if best_primary else None,
                "baseline_id": baseline_entity,
                "baseline_score": baseline_value,
                "delta_baseline": delta_baseline,
                "delta_baseline_pct": delta_baseline_pct,
                "best_external_id": best_external[0] if best_external else None,
                "best_external_score": best_external[1] if best_external else None,
                "delta_external": delta_external,
                "delta_external_pct": delta_external_pct,
                "sources": metric["sources"],
                "evidence_notes": metric.get("evidence_notes", ""),
            }
        )

    total_cells = len(metrics) * len(entities)
    blank_count = total_cells - numeric_count
    return {
        "coverage": coverage,
        "numeric_count": numeric_count,
        "blank_count": blank_count,
        "zero_count": zero_count,
        "total_cells": total_cells,
        "win_counts": dict(win_counts),
        "best_primary_counts": dict(best_primary_counts),
        "best_external_counts": dict(best_external_counts),
        "summary_rows": summary_rows,
    }


def fmt_value(value: float | int | None, unit: str = "") -> str:
    if value is None:
        return ""
    if unit.lower() == "elo" or abs(value) >= 1000:
        return f"{value:,.1f}"
    if abs(value - round(value)) < 0.05:
        return f"{value:.0f}"
    return f"{value:.1f}"


def short_label(value: str, width: int = 42) -> str:
    return value if len(value) <= width else value[: width - 1] + "..."


def style_axes(ax) -> None:
    for spine in ["top", "right", "left"]:
        ax.spines[spine].set_visible(False)
    ax.spines["bottom"].set_color("#D0D7DE")
    ax.tick_params(axis="y", length=0, labelsize=9)
    ax.tick_params(axis="x", colors="#5F6B7A", labelsize=8)
    ax.grid(axis="x", color="#E7EBF0", linewidth=0.8)
    ax.set_axisbelow(True)


def value_axis_bounds(values: list[float | int]) -> tuple[float, float, float]:
    max_v = max(values)
    min_v = min(values)
    span = max_v - min_v
    margin = span * 0.14 if span else (abs(max_v) * 0.12 if max_v else 1)
    left = min(0, min_v) - margin if min_v < 0 else 0
    right = max(0, max_v) + margin
    return left, right, margin


def add_value_labels(ax, bars, values: list[float | int], unit: str, margin: float) -> None:
    offset = margin * 0.10 if margin else 0.1
    for bar, value in zip(bars, values):
        if value >= 0:
            x = value + offset
            ha = "left"
        else:
            x = value - offset
            ha = "right"
        ax.text(
            x,
            bar.get_y() + bar.get_height() / 2,
            fmt_value(value, unit),
            va="center",
            ha=ha,
            fontsize=9,
            color="#1F2937",
        )


def save_model_hbar(plt, data, title: str, xlabel: str, path: Path, id_to_entity, height=5.7) -> None:
    import numpy as np

    rows = [(eid, value) for eid, value in data if isinstance(value, (int, float))]
    rows.sort(key=lambda x: x[1])
    labels = [id_to_entity[eid]["label"] for eid, _ in rows]
    values = [value for _, value in rows]
    colors = [id_to_entity[eid]["color"] for eid, _ in rows]
    fig, ax = plt.subplots(figsize=(9.7, height), facecolor="white")
    y = np.arange(len(labels))
    bars = ax.barh(y, values, color=colors, height=0.58)
    ax.set_yticks(y, [short_label(label, 42) for label in labels])
    style_axes(ax)
    left, right, margin = value_axis_bounds(values or [1])
    ax.set_xlim(left, right)
    add_value_labels(ax, bars, values, "", margin)
    ax.set_title(title, loc="left", fontsize=14, fontweight="bold", color="#17324D", pad=12)
    ax.set_xlabel(xlabel, fontsize=9, color="#5F6B7A")
    ax.text(0, -0.14, "Colors are fixed by entity identity.", transform=ax.transAxes, fontsize=8, color="#7A869A")
    plt.tight_layout(pad=1.25)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def save_metric_chart(plt, metric, entities, path: Path) -> None:
    import numpy as np

    rows = [
        (entity, metric["scores"].get(entity["id"]))
        for entity in entities
        if isinstance(metric["scores"].get(entity["id"]), (int, float))
    ]
    rows.sort(key=lambda item: item[1], reverse=metric["higher_is_better"])
    labels = [entity["label"] for entity, _ in rows]
    values = [value for _, value in rows]
    colors = [entity["color"] for entity, _ in rows]
    fig, ax = plt.subplots(figsize=(9.8, 4.9), facecolor="white")
    y = np.arange(len(labels))
    bars = ax.barh(y, values, color=colors, height=0.56)
    ax.set_yticks(y, [short_label(label, 35) for label in labels])
    ax.invert_yaxis()
    style_axes(ax)
    left, right, margin = value_axis_bounds(values)
    ax.set_xlim(left, right)
    add_value_labels(ax, bars, values, metric["unit"], margin)
    title = f"{metric['category']} - {metric['name']}"
    ax.set_title("\n".join(textwrap.wrap(title, 72)), loc="left", fontsize=13, fontweight="bold", color="#17324D", pad=12)
    ax.set_xlabel(metric["unit"], fontsize=9, color="#5F6B7A")
    direction = "higher is better" if metric["higher_is_better"] else "lower is better"
    ax.text(0, -0.16, f"Fixed entity colors; missing values are omitted; {direction}.", transform=ax.transAxes, fontsize=8, color="#7A869A")
    plt.tight_layout(pad=1.35)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def save_gap_chart(plt, summary_row, id_to_entity, path: Path) -> bool:
    import numpy as np

    items = []
    if summary_row["best_primary_id"]:
        items.append((summary_row["best_primary_id"], summary_row["best_primary_score"]))
    if summary_row["baseline_id"] and isinstance(summary_row["baseline_score"], (int, float)):
        items.append((summary_row["baseline_id"], summary_row["baseline_score"]))
    if summary_row["best_external_id"]:
        items.append((summary_row["best_external_id"], summary_row["best_external_score"]))
    dedup = {}
    for eid, value in items:
        if eid and isinstance(value, (int, float)):
            dedup[eid] = value
    if len(dedup) < 2:
        return False
    rows = [(id_to_entity[eid], value) for eid, value in dedup.items()]
    rows.sort(key=lambda item: item[1], reverse=summary_row["higher_is_better"])
    labels = [entity["label"] for entity, _ in rows]
    values = [value for _, value in rows]
    colors = [entity["color"] for entity, _ in rows]
    fig, ax = plt.subplots(figsize=(9.8, 3.2), facecolor="white")
    y = np.arange(len(labels))
    bars = ax.barh(y, values, color=colors, height=0.54)
    ax.set_yticks(y, [short_label(label, 36) for label in labels])
    ax.invert_yaxis()
    style_axes(ax)
    left, right, margin = value_axis_bounds(values)
    ax.set_xlim(left, right)
    add_value_labels(ax, bars, values, summary_row["unit"], margin)
    title = f"Comparison - {summary_row['category']} - {summary_row['metric']}"
    ax.set_title("\n".join(textwrap.wrap(title, 74)), loc="left", fontsize=12.5, fontweight="bold", color="#17324D", pad=10)
    ax.set_xlabel(summary_row["unit"], fontsize=9, color="#5F6B7A")
    ax.text(0, -0.20, "Bars show actual participating entities using their fixed theme colors.", transform=ax.transAxes, fontsize=8, color="#7A869A")
    plt.tight_layout(pad=1.30)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return True


def save_category_chart(plt, category: str, metrics: list[dict[str, Any]], entities, path: Path) -> None:
    import numpy as np

    fig_w = 9.9
    fig_h = max(4.8, len(metrics) * 3.18)
    fig, axes = plt.subplots(len(metrics), 1, figsize=(fig_w, fig_h), facecolor="white")
    if not isinstance(axes, np.ndarray):
        axes = np.array([axes])
    for ax, metric in zip(axes, metrics):
        rows = [
            (entity, metric["scores"].get(entity["id"]))
            for entity in entities
            if isinstance(metric["scores"].get(entity["id"]), (int, float))
        ]
        rows.sort(key=lambda item: item[1], reverse=metric["higher_is_better"])
        labels = [entity["label"] for entity, _ in rows]
        values = [value for _, value in rows]
        colors = [entity["color"] for entity, _ in rows]
        bars = ax.barh(np.arange(len(labels)), values, color=colors, height=0.54)
        ax.set_yticks(np.arange(len(labels)), [short_label(label, 35) for label in labels])
        ax.invert_yaxis()
        style_axes(ax)
        left, right, margin = value_axis_bounds(values)
        ax.set_xlim(left, right)
        add_value_labels(ax, bars, values, metric["unit"], margin)
        title = textwrap.shorten(metric["name"], width=74, placeholder="...") + f" ({metric['unit']})"
        ax.set_title(title, loc="left", fontsize=10.5, fontweight="bold", color="#17324D", pad=8)
    fig.suptitle(category, fontsize=16, fontweight="bold", color="#17324D", x=0.01, ha="left", y=0.998)
    plt.tight_layout(rect=[0, 0, 1, 0.985], h_pad=1.45)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def build_workbook(spec: dict[str, Any], output: Path) -> dict[str, Any]:
    require_dependencies()
    plt = setup_matplotlib()
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    entities = spec["entities"]
    id_to_entity = {entity["id"]: entity for entity in entities}
    metrics = spec["metrics"]
    stats = compute_model_stats(spec)
    assets = Path(tempfile.mkdtemp(prefix="bar_chart_research_"))

    image_paths = []
    coverage_img = assets / "coverage.png"
    wins_img = assets / "wins.png"
    save_model_hbar(
        plt,
        [(entity["id"], stats["coverage"][entity["id"]]) for entity in entities],
        "Data coverage by entity",
        "Published score count",
        coverage_img,
        id_to_entity,
        height=5.7,
    )
    image_paths.append(coverage_img)
    save_model_hbar(
        plt,
        [(entity["id"], stats["win_counts"].get(entity["id"], 0)) for entity in entities],
        "Overall winner count",
        "Count",
        wins_img,
        id_to_entity,
        height=5.7,
    )
    image_paths.append(wins_img)

    metric_images = []
    for idx, metric in enumerate(metrics, 1):
        path = assets / f"metric_{idx:03d}.png"
        save_metric_chart(plt, metric, entities, path)
        metric_images.append(path)
        image_paths.append(path)

    gap_images = []
    for idx, row in enumerate(stats["summary_rows"], 1):
        path = assets / f"gap_{idx:03d}.png"
        if save_gap_chart(plt, row, id_to_entity, path):
            gap_images.append(path)
            image_paths.append(path)

    category_images = []
    category_order = []
    for metric in metrics:
        if metric["category"] not in category_order:
            category_order.append(metric["category"])
    for idx, category in enumerate(category_order, 1):
        path = assets / f"category_{idx:03d}.png"
        category_metrics = [metric for metric in metrics if metric["category"] == category]
        save_category_chart(plt, category, category_metrics, entities, path)
        category_images.append((category, path))
        image_paths.append(path)

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="17324D")
    soft_fill = PatternFill("solid", fgColor="F3F7FB")
    pos_fill = PatternFill("solid", fgColor="E2F0D9")
    neg_fill = PatternFill("solid", fgColor="FCE4D6")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=20, bold=True, color="17324D")
    section_font = Font(size=13, bold=True, color="17324D")
    small_font = Font(color="666666", size=10)
    link_font = Font(color="0563C1", underline="single")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def add_image(ws, path: Path, anchor: str, scale: float = 0.66) -> int:
        img = XLImage(str(path))
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        ws.add_image(img, anchor)
        return img.height

    def prepare_chart_sheet(ws, title: str, subtitle: str) -> None:
        ws.sheet_view.showGridLines = False
        ws["A1"] = title
        ws["A1"].font = title_font
        ws["A2"] = subtitle
        ws["A2"].font = small_font
        ws.merge_cells("A1:L1")
        ws.merge_cells("A2:L2")
        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 12
        for row in range(1, 1500):
            ws.row_dimensions[row].height = 18

    ws = wb.active
    ws.title = "00_阅读说明"
    readme_rows = [
        ["File", spec["title"]],
        ["Generated", spec["generated_at"]],
        ["Source note", spec.get("source_note", "")],
        ["Missing policy", "Missing values stay blank/null and are not counted as zero."],
        ["Entity cells", stats["total_cells"]],
        ["Numeric cells", stats["numeric_count"]],
        ["Blank cells", stats["blank_count"]],
        ["Zero-valued cells", stats["zero_count"]],
    ]
    for row in readme_rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 110

    ws = wb.create_sheet("01_总览")
    prepare_chart_sheet(ws, spec["title"], "Entity colors are fixed globally. Missing values are omitted, not converted to zero.")
    cards = [
        ("Metrics", str(len(metrics))),
        ("Entities", str(len(entities))),
        ("Numeric cells", str(stats["numeric_count"])),
        ("Blank cells", str(stats["blank_count"])),
        ("Zero cells", str(stats["zero_count"])),
        ("Chart images", str(len(image_paths))),
    ]
    for idx, (label, value) in enumerate(cards):
        row = 4 + (idx // 2) * 2
        col = 1 + (idx % 2) * 4
        ws.cell(row, col).value = label
        ws.cell(row + 1, col).value = value
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 2)
        for c in range(col, col + 3):
            ws.cell(row, c).fill = header_fill
            ws.cell(row, c).font = white_font
            ws.cell(row, c).alignment = center
            ws.cell(row, c).border = border
            ws.cell(row + 1, c).fill = soft_fill
            ws.cell(row + 1, c).font = Font(size=15, bold=True, color="17324D")
            ws.cell(row + 1, c).alignment = center
            ws.cell(row + 1, c).border = border
    row_cursor = 12
    for image in [coverage_img, wins_img]:
        height = add_image(ws, image, f"A{row_cursor}", scale=0.68)
        row_cursor += math.ceil(height / 18) + 5

    ws = wb.create_sheet("02_所有柱状图")
    prepare_chart_sheet(ws, "All metric bar charts", "Each chart uses fixed entity theme colors.")
    ws["A4"] = "Color legend"
    ws["A4"].font = section_font
    for idx, entity in enumerate(entities):
        row = 5 + idx
        ws.cell(row, 1).value = entity["label"]
        ws.cell(row, 2).value = entity["group"]
        ws.cell(row, 3).value = entity["color"]
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=sheet_color(entity["color"]))
        ws.cell(row, 1).font = white_font
        for c in [1, 2, 3]:
            ws.cell(row, c).border = border
            ws.cell(row, c).alignment = center
    row_cursor = 5 + len(entities) + 3
    for image in metric_images:
        height = add_image(ws, image, f"A{row_cursor}", scale=0.65)
        row_cursor += math.ceil(height / 18) + 4

    if gap_images:
        ws = wb.create_sheet("03_差距对比图")
        prepare_chart_sheet(ws, "Comparison charts", "Bars show actual participating entities with their fixed theme colors.")
        row_cursor = 4
        for image in gap_images:
            height = add_image(ws, image, f"A{row_cursor}", scale=0.66)
            row_cursor += math.ceil(height / 18) + 4

    ws = wb.create_sheet("04_按领域")
    prepare_chart_sheet(ws, "Category charts", "One composite chart per category.")
    row_cursor = 4
    for category, image in category_images:
        ws.cell(row_cursor, 1).value = category
        ws.cell(row_cursor, 1).font = section_font
        height = add_image(ws, image, f"A{row_cursor + 1}", scale=0.63)
        row_cursor += math.ceil(height / 18) + 6

    ws = wb.create_sheet("05_颜色标准")
    ws.append(["Entity", "Group", "Fixed color HEX", "Swatch", "Rule"])
    for entity in entities:
        ws.append([entity["label"], entity["group"], entity["color"], "", "Use this color for this entity in every chart/table."])
    for row in range(2, ws.max_row + 1):
        color = ws.cell(row, 3).value
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=sheet_color(color))
        ws.cell(row, 1).font = white_font
        ws.cell(row, 4).fill = PatternFill("solid", fgColor=sheet_color(color))
        ws.cell(row, 4).value = "      "

    ws = wb.create_sheet("06_缺失值审计")
    ws.append(["Entity", "Group", "Numeric scores", "Missing scores", "Coverage", "Missing counted as zero?"])
    for entity in entities:
        numeric = stats["coverage"][entity["id"]]
        missing = len(metrics) - numeric
        ws.append([entity["label"], entity["group"], numeric, missing, numeric / len(metrics), "No"])
    for row in range(2, ws.max_row + 1):
        label = ws.cell(row, 1).value
        entity = next(e for e in entities if e["label"] == label)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=sheet_color(entity["color"]))
        ws.cell(row, 1).font = white_font
        ws.cell(row, 5).number_format = "0.0%"

    ws = wb.create_sheet("07_差距摘要")
    summary_headers = [
        "Category",
        "Metric",
        "Unit",
        "Higher is better",
        "Winner",
        "Best primary",
        "Best primary score",
        "Baseline",
        "Baseline score",
        "Delta vs baseline",
        "Delta vs baseline %",
        "Best external",
        "Best external score",
        "Delta vs external",
        "Delta vs external %",
        "Evidence notes",
        "Sources",
    ]
    ws.append(summary_headers)
    for row in stats["summary_rows"]:
        ws.append(
            [
                row["category"],
                row["metric"],
                row["unit"],
                row["higher_is_better"],
                row["winner"],
                id_to_entity[row["best_primary_id"]]["label"] if row["best_primary_id"] else "",
                row["best_primary_score"],
                id_to_entity[row["baseline_id"]]["label"] if row["baseline_id"] in id_to_entity else "",
                row["baseline_score"],
                row["delta_baseline"],
                row["delta_baseline_pct"],
                id_to_entity[row["best_external_id"]]["label"] if row["best_external_id"] else "",
                row["best_external_score"],
                row["delta_external"],
                row["delta_external_pct"],
                row["evidence_notes"],
                "\n".join(row["sources"]),
            ]
        )
    for row in range(2, ws.max_row + 1):
        for col in [10, 11, 14, 15]:
            value = ws.cell(row, col).value
            if isinstance(value, (int, float)):
                ws.cell(row, col).fill = pos_fill if value >= 0 else neg_fill

    ws = wb.create_sheet("08_原始数据")
    raw_headers = ["Category", "Metric", "Unit", "Higher is better"] + [entity["label"] for entity in entities] + ["Evidence notes", "Sources"]
    ws.append(raw_headers)
    for metric in metrics:
        ws.append(
            [
                metric["category"],
                metric["name"],
                metric["unit"],
                metric["higher_is_better"],
                *[metric["scores"].get(entity["id"]) for entity in entities],
                metric.get("evidence_notes", ""),
                "\n".join(metric["sources"]),
            ]
        )
    for col, entity in enumerate(entities, start=5):
        ws.cell(1, col).fill = PatternFill("solid", fgColor=sheet_color(entity["color"]))
        ws.cell(1, col).font = white_font

    ws = wb.create_sheet("09_来源")
    ws.append(["Source", "Notes"])
    seen_sources = []
    source_notes = {}
    for metric in metrics:
        for source in metric["sources"]:
            if source not in seen_sources:
                seen_sources.append(source)
            if metric.get("evidence_notes"):
                source_notes.setdefault(source, set()).add(metric["evidence_notes"])
    for source in seen_sources:
        ws.append([source, "\n".join(sorted(source_notes.get(source, set())))])
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, str) and value.startswith(("http://", "https://")):
            ws.cell(row, 1).hyperlink = value
            ws.cell(row, 1).font = link_font

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        max_row, max_col = ws.max_row, ws.max_column
        if not ws.title.startswith(("01_", "02_", "03_", "04_")):
            for cell in ws[1]:
                if cell.fill.fill_type is None:
                    cell.fill = header_fill
                    cell.font = white_font
                cell.alignment = center
                cell.border = border
            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.alignment = wrap
                    cell.border = border
        if ws.title.startswith(("05_", "06_", "07_", "08_", "09_")):
            try:
                table = Table(displayName="T" + str(wb.worksheets.index(ws) + 1), ref=ws.dimensions)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                ws.add_table(table)
            except Exception:
                ws.auto_filter.ref = ws.dimensions
        if not ws.title.startswith(("01_", "02_", "03_", "04_")):
            for col in range(1, max_col + 1):
                letter = get_column_letter(col)
                max_len = 0
                for row in range(1, min(max_row, 80) + 1):
                    value = ws.cell(row, col).value
                    if value is not None:
                        max_len = max(max_len, len(str(value)))
                ws.column_dimensions[letter].width = min(max(max_len * 1.1, 11), 56)
            for row in range(1, max_row + 1):
                ws.row_dimensions[row].height = 24 if row == 1 else 34

    wb.properties.title = spec["title"]
    wb.properties.subject = "Source-backed bar chart research workbook"
    wb.properties.creator = "bar-chart-research skill"
    wb.properties.created = datetime.now()
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    check = load_workbook(output, read_only=False, data_only=True)
    embedded_images = sum(len(getattr(ws, "_images", [])) for ws in check.worksheets)
    result = {
        "output": str(output),
        "sheets": check.sheetnames,
        "metrics": len(metrics),
        "entities": len(entities),
        "total_cells": stats["total_cells"],
        "numeric_cells": stats["numeric_count"],
        "blank_cells": stats["blank_count"],
        "zero_cells": stats["zero_count"],
        "embedded_images": embedded_images,
        "size_bytes": output.stat().st_size,
    }
    shutil.rmtree(assets, ignore_errors=True)
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Create a bar-chart research Excel workbook.")
    parser.add_argument("--input", type=Path, help="Input JSON spec.")
    parser.add_argument("--output", type=Path, required=True, help="Output .xlsx path.")
    parser.add_argument("--self-test", action="store_true", help="Generate a workbook from synthetic test data.")
    args = parser.parse_args(argv)

    spec = normalize_spec(load_spec(args.input, args.self_test))
    result = build_workbook(spec, args.output)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
